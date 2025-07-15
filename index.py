from flask import Flask, request, send_file, render_template_string
import fitz  # PyMuPDF
import json
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO

app = Flask(__name__)
HTML_FORM = '''
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Bank PDF to Excel Converter</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
  <style>
    * {
      margin: 0;
      padding: 0;
      box-sizing: border-box;
    }

    body {
      height: 100vh;
      background: radial-gradient(ellipse at top, #0f2027, #203a43, #2c5364);
      display: flex;
      justify-content: center;
      align-items: center;
      font-family: 'Inter', sans-serif;
      overflow: hidden;
      position: relative;
    }

    /* Animated Background Stars */
    .stars {
      position: absolute;
      top: 0;
      left: 0;
      width: 100%;
      height: 100%;
      pointer-events: none;
    }

    .star {
      position: absolute;
      width: 2px;
      height: 2px;
      background: #fff;
      border-radius: 50%;
      animation: twinkle 3s infinite;
    }

    .star:nth-child(odd) {
      animation-delay: -1s;
    }

    .star:nth-child(3n) {
      animation-delay: -2s;
    }

    @keyframes twinkle {
      0%, 100% { opacity: 0.3; transform: scale(1); }
      50% { opacity: 1; transform: scale(1.2); }
    }

    /* Floating particles */
    .particle {
      position: absolute;
      width: 4px;
      height: 4px;
      background: rgba(255, 255, 255, 0.4);
      border-radius: 50%;
      animation: float 6s infinite linear;
    }

    @keyframes float {
      0% { transform: translateY(100vh) rotate(0deg); opacity: 0; }
      10% { opacity: 1; }
      90% { opacity: 1; }
      100% { transform: translateY(-100px) rotate(360deg); opacity: 0; }
    }

    .container {
      background: rgba(255, 255, 255, 0.08);
      border: 1px solid rgba(255, 255, 255, 0.18);
      border-radius: 24px;
      padding: 48px 40px;
      width: 420px;
      max-width: 90vw;
      backdrop-filter: blur(16px);
      box-shadow: 
        0 8px 32px 0 rgba(0, 0, 0, 0.37),
        0 0 0 1px rgba(255, 255, 255, 0.05);
      text-align: center;
      color: white;
      position: relative;
      z-index: 10;
      transition: transform 0.3s ease;
    }

    .container:hover {
      transform: translateY(-2px);
    }

    .logo {
      width: 48px;
      height: 48px;
      background: linear-gradient(135deg, #00c9ff, #92fe9d);
      border-radius: 12px;
      display: flex;
      align-items: center;
      justify-content: center;
      margin: 0 auto 24px;
      position: relative;
    }

    .logo::before {
      content: '📄';
      font-size: 24px;
    }

    h2 {
      margin-bottom: 8px;
      font-size: 28px;
      font-weight: 600;
      color: #fff;
      background: linear-gradient(135deg, #fff, #e0e7ff);
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
    }

    .subtitle {
      color: rgba(255, 255, 255, 0.7);
      font-size: 14px;
      margin-bottom: 32px;
      font-weight: 400;
    }

    .file-input-wrapper {
      position: relative;
      margin-bottom: 24px;
    }

    .file-input-label {
      display: block;
      font-size: 14px;
      font-weight: 500;
      color: rgba(255, 255, 255, 0.9);
      margin-bottom: 8px;
      text-align: left;
    }

    input[type="file"] {
      background: rgba(255, 255, 255, 0.05);
      color: #fff;
      border: 1px solid rgba(255, 255, 255, 0.2);
      border-radius: 12px;
      padding: 16px;
      width: 100%;
      cursor: pointer;
      font-family: 'Inter', sans-serif;
      font-size: 14px;
      transition: all 0.3s ease;
    }

    input[type="file"]:hover {
      background: rgba(255, 255, 255, 0.08);
      border-color: rgba(255, 255, 255, 0.3);
    }

    input[type="file"]:focus {
      outline: none;
      border-color: #00c9ff;
      box-shadow: 0 0 0 3px rgba(0, 201, 255, 0.1);
    }

    input[type="file"]::-webkit-file-upload-button {
      background: linear-gradient(135deg, #2980b9, #3498db);
      color: white;
      padding: 8px 16px;
      border: none;
      border-radius: 8px;
      cursor: pointer;
      font-weight: 500;
      font-size: 12px;
      margin-right: 12px;
      transition: all 0.3s ease;
    }

    input[type="file"]::-webkit-file-upload-button:hover {
      background: linear-gradient(135deg, #3498db, #5dade2);
      transform: translateY(-1px);
    }

    .file-status {
      margin-top: 8px;
      font-size: 12px;
      color: #92fe9d;
      text-align: left;
      opacity: 0;
      transform: translateY(-10px);
      transition: all 0.3s ease;
    }

    .file-status.show {
      opacity: 1;
      transform: translateY(0);
    }

    button {
      background: linear-gradient(135deg, #00c9ff, #92fe9d);
      border: none;
      border-radius: 12px;
      padding: 16px 24px;
      width: 100%;
      font-size: 16px;
      font-weight: 600;
      color: #000;
      cursor: pointer;
      transition: all 0.3s ease;
      font-family: 'Inter', sans-serif;
      position: relative;
      overflow: hidden;
    }

    button:hover {
      transform: translateY(-2px);
      box-shadow: 0 8px 25px rgba(0, 201, 255, 0.3);
    }

    button:active {
      transform: translateY(0);
    }

    button:disabled {
      background: rgba(255, 255, 255, 0.1);
      color: rgba(255, 255, 255, 0.5);
      cursor: not-allowed;
      transform: none;
      box-shadow: none;
    }

    .loading {
      display: none;
      align-items: center;
      justify-content: center;
      gap: 8px;
    }

    .spinner {
      width: 16px;
      height: 16px;
      border: 2px solid rgba(0, 0, 0, 0.3);
      border-top: 2px solid #000;
      border-radius: 50%;
      animation: spin 1s linear infinite;
    }

    @keyframes spin {
      0% { transform: rotate(0deg); }
      100% { transform: rotate(360deg); }
    }

    .features {
      margin-top: 32px;
      padding-top: 24px;
      border-top: 1px solid rgba(255, 255, 255, 0.1);
    }

    .features-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 24px;
      margin-bottom: 24px;
    }

    .feature {
      text-align: center;
    }

    .feature-title {
      font-weight: 600;
      font-size: 18px;
      color: #00c9ff;
      margin-bottom: 4px;
    }

    .feature-desc {
      font-size: 12px;
      color: rgba(255, 255, 255, 0.6);
    }

    .footer {
      font-size: 12px;
      color: rgba(255, 255, 255, 0.7);
      padding-top: 16px;
      border-top: 1px solid rgba(255, 255, 255, 0.1);
    }

    .support-text {
      margin-top: 24px;
      font-size: 14px;
      color: rgba(255, 255, 255, 0.6);
    }

    /* Responsive Design */
    @media (max-width: 480px) {
      .container {
        padding: 32px 24px;
        margin: 20px;
      }
      
      h2 {
        font-size: 24px;
      }
      
      .features-grid {
        grid-template-columns: 1fr;
        gap: 16px;
      }
    }
  </style>
</head>
<body>
  <!-- Animated Background -->
  <div class="stars" id="stars"></div>

  <div class="container">
    <div class="logo"></div>
    <h2>PDF Converter</h2>
    <p class="subtitle">Convert Bank PDFs to Excel instantly</p>
    
    <form method="POST" action="/convert" enctype="multipart/form-data" id="convertForm">
      <div class="file-input-wrapper">
        <label class="file-input-label">Select PDF File</label>
        <input type="file" name="pdf_file" accept=".pdf" required id="fileInput">
        <div class="file-status" id="fileStatus"></div>
      </div>
      
      <button type="submit" id="convertBtn">
        <span class="btn-text">Convert & Download</span>
        <div class="loading">
          <div class="spinner"></div>
          <span>Converting...</span>
        </div>
      </button>
    </form>

    <div class="features">
      <div class="features-grid">
        <div class="feature">
          <div class="feature-title">Fast</div>
          <div class="feature-desc">Instant conversion</div>
        </div>
        <div class="feature">
          <div class="feature-title">Secure</div>
          <div class="feature-desc">Your data is safe</div>
        </div>
      </div>
      
      <div class="footer">MADE FOR ARUNKUMAR 📉</div>
    </div>

    <div class="support-text">
      Supported formats: PDF → Excel (XLSX)
    </div>
  </div>

  <script>
    // Create animated stars
    function createStars() {
      const starsContainer = document.getElementById('stars');
      const starCount = 100;
      
      for (let i = 0; i < starCount; i++) {
        const star = document.createElement('div');
        star.className = 'star';
        star.style.left = Math.random() * 100 + '%';
        star.style.top = Math.random() * 100 + '%';
        star.style.animationDelay = Math.random() * 3 + 's';
        star.style.animationDuration = (2 + Math.random() * 2) + 's';
        starsContainer.appendChild(star);
      }
    }

    // Create floating particles
    function createParticles() {
      setInterval(() => {
        const particle = document.createElement('div');
        particle.className = 'particle';
        particle.style.left = Math.random() * 100 + '%';
        particle.style.animationDuration = (4 + Math.random() * 2) + 's';
        document.body.appendChild(particle);
        
        setTimeout(() => {
          if (document.body.contains(particle)) {
            document.body.removeChild(particle);
          }
        }, 6000);
      }, 500);
    }

    // File input handler
    document.getElementById('fileInput').addEventListener('change', function(e) {
      const fileStatus = document.getElementById('fileStatus');
      if (e.target.files && e.target.files[0]) {
        fileStatus.textContent = `✓ ${e.target.files[0].name} selected`;
        fileStatus.classList.add('show');
      } else {
        fileStatus.classList.remove('show');
      }
    });

    // Form submission handler
    document.getElementById('convertForm').addEventListener('submit', function(e) {
      const btn = document.getElementById('convertBtn');
      const btnText = btn.querySelector('.btn-text');
      const loading = btn.querySelector('.loading');
      
      // Show loading state
      btn.disabled = true;
      btnText.style.display = 'none';
      loading.style.display = 'flex';
      
      // The form will submit to your Flask backend
      // The loading state will be reset when the page reloads or the download starts
    });

    // Initialize animations
    createStars();
    createParticles();
  </script>
</body>
</html>
'''

@app.route("/")
def index():
    return render_template_string(HTML_FORM)

@app.route("/convert", methods=["POST"])
def convert_pdf_to_excel():
    uploaded_file = request.files["pdf_file"]
    if not uploaded_file:
        return "No file uploaded"

    pdf_bytes = uploaded_file.read()
    doc = fitz.open("pdf", pdf_bytes)

    lines = []
    for page in doc:
        lines.extend(page.get_text().split("\n"))

    # Detect opening balance
    opening_balance = None
    amount_pattern = re.compile(r"^\d{1,3}(?:,\d{3})*(?:\.\d{2})$")
    for i in range(len(lines)):
        if "Opening Balance" in lines[i]:
            for j in range(i+1, i+4):
                if amount_pattern.match(lines[j].strip()):
                    opening_balance = float(lines[j].replace(",", ""))
                    break
            break

    if opening_balance is None:
        return "Opening Balance not found."

    transactions = []
    i = 0
    date_pattern = re.compile(r"\d{2}-\d{2}-\d{4}")
    previous_balance = opening_balance

    while i < len(lines):
        line = lines[i].strip()
        if date_pattern.match(line):
            date = line
            i += 1
            particulars = []

            while i < len(lines) and not lines[i].startswith("Chq:") and not amount_pattern.match(lines[i].strip()):
                particulars.append(lines[i].strip())
                i += 1

            if i < len(lines) and lines[i].startswith("Chq:"):
                particulars.append(lines[i].strip())
                i += 1

            amounts = []
            while i < len(lines) and len(amounts) < 2:
                amt_line = lines[i].strip()
                if amount_pattern.match(amt_line):
                    amounts.append(amt_line)
                i += 1

            deposit, withdrawal, balance = "", "", ""
            if len(amounts) == 2:
                amount_val = float(amounts[0].replace(",", ""))
                balance_val = float(amounts[1].replace(",", ""))
                balance = amounts[1]

                if balance_val > previous_balance:
                    deposit = amounts[0]
                elif balance_val < previous_balance:
                    withdrawal = amounts[0]

                previous_balance = balance_val

            elif len(amounts) == 1:
                balance = amounts[0]
                previous_balance = float(balance.replace(",", ""))

            transactions.append({
                "date": date,
                "particulars": " ".join(particulars),
                "deposit": deposit,
                "withdrawal": withdrawal,
                "balance": balance
            })
        else:
            i += 1

    # Create Excel
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Transactions"

    headers = ["Date", "Particulars", "Deposit", "Withdrawal", "Balance"]
    ws.append(headers)

    for tx in transactions:
        ws.append([tx["date"], tx["particulars"], tx["deposit"], tx["withdrawal"], tx["balance"]])

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="converted_statement.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=10000)
